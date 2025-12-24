#!/usr/bin/env python3
"""
Activity 02: Data Cleansing
Exports foodraw.txt to foodraw.xlsx with cleaned data
- Columns: id, name, link, ingredients_raw, allergens_raw, ingredients, allergensraw
- Ingredients and allergens cleaned using EXCEL FORMULAS (lowercase, alphabets and commas only)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def build_cleaning_formula(cell_ref):
    """
    Build an Excel formula that:
    1. Converts text to lowercase
    2. Removes numbers and special characters (keeps only letters, spaces, commas)

    Uses nested SUBSTITUTE functions to remove unwanted characters
    """
    # Characters to remove (numbers and common special characters)
    chars_to_remove = [
        "0", "1", "2", "3", "4", "5", "6", "7", "8", "9",
        "(", ")", "[", "]", "{", "}",
        ".", ":", ";", "!", "?",
        "/", "\\", "-", "_", "*", "#", "@",
        "%", "&", "+", "=", "<", ">",
        "'", '"', "`", "~", "^", "|"
    ]

    # Start with LOWER() to convert to lowercase
    formula = f"LOWER({cell_ref})"

    # Wrap with nested SUBSTITUTE to remove each unwanted character
    for char in chars_to_remove:
        # Escape double quotes for Excel formula
        if char == '"':
            formula = f'SUBSTITUTE({formula},CHAR(34),"")'
        else:
            formula = f'SUBSTITUTE({formula},"{char}","")'

    # Final formula with TRIM to clean up extra spaces
    formula = f"=TRIM({formula})"

    return formula


def parse_foodraw(filename="foodraw.txt"):
    """Parse the foodraw.txt file"""
    products = []

    with open(filename, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            # Split by semicolon
            parts = line.split(";")

            if len(parts) >= 5:
                product = {
                    "id": parts[0].strip(),
                    "name": parts[1].strip(),
                    "ingredients": parts[2].strip(),
                    "allergensraw": parts[3].strip(),
                    "link": parts[4].strip()
                }
                products.append(product)
            elif len(parts) == 4:
                # Missing allergens
                product = {
                    "id": parts[0].strip(),
                    "name": parts[1].strip(),
                    "ingredients": parts[2].strip(),
                    "allergensraw": "",
                    "link": parts[3].strip()
                }
                products.append(product)

    return products


def create_excel(products, filename="foodraw.xlsx"):
    """Create Excel file with raw data and Excel formulas for cleaning"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Food Data"

    # Define headers
    # Columns A-C: id, name, link
    # Columns D-E: raw data (ingredients_raw, allergens_raw)
    # Columns F-G: cleaned data using Excel formulas (ingredients, allergensraw)
    headers = ["id", "name", "link", "ingredients_raw", "allergens_raw", "ingredients", "allergensraw"]

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    raw_header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange for raw
    formula_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green for formula
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers with different colors
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        # Color coding: blue for basic, orange for raw, green for formula-cleaned
        if col <= 3:
            cell.fill = header_fill  # Blue
        elif col <= 5:
            cell.fill = raw_header_fill  # Orange (raw data)
        else:
            cell.fill = formula_header_fill  # Green (formula cleaned)

    # Write data
    for row_idx, product in enumerate(products, 2):
        # Column A: id (as-is)
        ws.cell(row=row_idx, column=1, value=product["id"])

        # Column B: name (as-is)
        ws.cell(row=row_idx, column=2, value=product["name"])

        # Column C: link (as-is)
        ws.cell(row=row_idx, column=3, value=product["link"])

        # Column D: ingredients_raw (original raw data)
        ws.cell(row=row_idx, column=4, value=product["ingredients"])

        # Column E: allergens_raw (original raw data)
        ws.cell(row=row_idx, column=5, value=product["allergensraw"])

        # Column F: ingredients (EXCEL FORMULA - references D column)
        ingredients_formula = build_cleaning_formula(f"D{row_idx}")
        ws.cell(row=row_idx, column=6, value=ingredients_formula)

        # Column G: allergensraw (EXCEL FORMULA - references E column)
        allergens_formula = build_cleaning_formula(f"E{row_idx}")
        ws.cell(row=row_idx, column=7, value=allergens_formula)

        # Apply border to all cells
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 8   # id
    ws.column_dimensions['B'].width = 40  # name
    ws.column_dimensions['C'].width = 50  # link
    ws.column_dimensions['D'].width = 80  # ingredients_raw
    ws.column_dimensions['E'].width = 40  # allergens_raw
    ws.column_dimensions['F'].width = 80  # ingredients (formula)
    ws.column_dimensions['G'].width = 40  # allergensraw (formula)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(filename)
    print(f"Saved Excel file: {filename}")
    print(f"  - Columns D & E contain RAW data")
    print(f"  - Columns F & G contain EXCEL FORMULAS that clean the data")

    return wb


def main():
    print("=" * 60)
    print("Activity 02: Data Cleansing")
    print("=" * 60)

    # Parse the foodraw.txt
    print("\nParsing foodraw.txt...")
    products = parse_foodraw("foodraw.txt")
    print(f"Found {len(products)} products")

    # Create Excel file
    print("\nCreating Excel file with Excel formulas for cleaning...")
    create_excel(products, "foodraw.xlsx")

    # Summary
    with_allergens = sum(1 for p in products if p["allergensraw"].strip())
    without_allergens = len(products) - with_allergens

    print("\n" + "=" * 60)
    print("Summary:")
    print(f"  Total products: {len(products)}")
    print(f"  With allergens: {with_allergens}")
    print(f"  Without allergens: {without_allergens}")
    print("=" * 60)

    # Show Excel structure
    print("\nExcel file structure:")
    print("  Column A: id")
    print("  Column B: name")
    print("  Column C: link")
    print("  Column D: ingredients_raw (original data)")
    print("  Column E: allergens_raw (original data)")
    print("  Column F: ingredients (EXCEL FORMULA - cleaned)")
    print("  Column G: allergensraw (EXCEL FORMULA - cleaned)")
    print("\nThe formulas in columns F & G will:")
    print("  - Convert text to lowercase using LOWER()")
    print("  - Remove numbers and special characters using nested SUBSTITUTE()")
    print("  - Keep only alphabets, spaces, and commas")


if __name__ == "__main__":
    main()