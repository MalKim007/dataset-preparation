#!/usr/bin/env python3
"""
Activity 03: Data Mapping
Maps allergens to the nine common allergens:
1. Milk
2. Egg
3. Peanut
4. Tree nut
5. Wheat
6. Soy
7. Fish
8. Shellfish
9. Sesame

Creates foodpreprocessed.xlsx with allergensmapped column
"""

import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Mapping rules for nine common allergens
ALLERGEN_MAPPINGS = {
    "milk": [
        "milk", "dairy", "lactose", "casein", "whey", "butter", "cream", "cheese",
        "yogurt", "yoghurt", "ghee", "paneer", "curd", "lactalbumin", "lactoglobulin",
        "milk protein", "milk powder", "skimmed milk", "whole milk", "condensed milk",
        "evaporated milk", "buttermilk", "milk fat", "milk solids", "lait", "milch"
    ],
    "egg": [
        "egg", "eggs", "albumin", "albumen", "globulin", "lysozyme", "mayonnaise",
        "meringue", "ovalbumin", "ovomucin", "ovomucoid", "ovovitellin", "egg white",
        "egg yolk", "egg protein", "dried egg", "egg powder", "whole egg",
        "pasteurized egg", "liquid egg"
    ],
    "peanut": [
        "peanut", "peanuts", "groundnut", "groundnuts", "arachis", "monkey nut",
        "earth nut", "beer nut", "peanut oil", "peanut butter", "peanut flour",
        "arachis hypogaea", "goober"
    ],
    "tree nut": [
        "almond", "almonds", "cashew", "cashews", "walnut", "walnuts", "pecan", "pecans",
        "pistachio", "pistachios", "hazelnut", "hazelnuts", "filbert", "filberts",
        "macadamia", "brazil nut", "brazil nuts", "chestnut", "chestnuts", "pine nut",
        "pine nuts", "praline", "marzipan", "nougat", "gianduja", "tree nut", "tree nuts",
        "nut", "nuts", "mixed nuts", "coconut"  # Note: coconut is sometimes classified separately
    ],
    "wheat": [
        "wheat", "gluten", "flour", "bread", "breadcrumbs", "bulgur", "couscous",
        "durum", "einkorn", "emmer", "farina", "kamut", "semolina", "spelt", "triticale",
        "wheat germ", "wheat bran", "wheat starch", "wheat protein", "wheat flour",
        "whole wheat", "enriched flour", "all purpose flour", "bread flour", "cake flour",
        "seitan", "vital wheat gluten", "modified wheat starch"
    ],
    "soy": [
        "soy", "soya", "soybean", "soybeans", "edamame", "miso", "tempeh", "tofu",
        "soy sauce", "soy protein", "soy lecithin", "soy flour", "soy milk", "soy oil",
        "textured vegetable protein", "tvp", "hydrolyzed soy", "soy isolate",
        "soy concentrate", "soya bean", "soja"
    ],
    "fish": [
        "fish", "salmon", "tuna", "cod", "haddock", "anchovy", "anchovies", "sardine",
        "sardines", "mackerel", "herring", "trout", "bass", "tilapia", "pollock",
        "catfish", "perch", "pike", "carp", "halibut", "flounder", "sole", "snapper",
        "grouper", "swordfish", "mahi", "fish sauce", "fish oil", "fish protein",
        "fish gelatin", "omega 3", "dha", "epa"
    ],
    "shellfish": [
        "shellfish", "shrimp", "prawn", "prawns", "crab", "lobster", "crayfish",
        "crawfish", "langoustine", "scallop", "scallops", "clam", "clams", "mussel",
        "mussels", "oyster", "oysters", "squid", "calamari", "octopus", "snail",
        "escargot", "abalone", "cockle", "crustacean", "crustaceans", "mollusc",
        "mollusk", "molluscs", "mollusks"
    ],
    "sesame": [
        "sesame", "sesame seed", "sesame seeds", "tahini", "tahina", "halvah", "halva",
        "hummus", "houmous", "sesame oil", "sesame paste", "sesame flour", "goma",
        "til", "benne seed", "sesamum"
    ]
}


def map_to_common_allergens(allergen_text):
    """
    Map allergen text to the nine common allergens
    Returns a comma-separated string of matched allergens in lowercase
    """
    if not allergen_text or allergen_text.strip() == "":
        return ""

    allergen_lower = allergen_text.lower()
    found_allergens = set()

    for common_allergen, keywords in ALLERGEN_MAPPINGS.items():
        for keyword in keywords:
            # Use word boundary matching to avoid partial matches
            pattern = r'\b' + re.escape(keyword) + r'\b'
            if re.search(pattern, allergen_lower):
                found_allergens.add(common_allergen)
                break

    # Sort alphabetically for consistency
    sorted_allergens = sorted(found_allergens)
    return ", ".join(sorted_allergens)


def process_excel(input_file="foodraw.xlsx", output_file="foodpreprocessed.xlsx"):
    """
    Process the Excel file and add allergensmapped column
    """
    # Load the existing workbook
    wb = load_workbook(input_file)
    ws = wb.active

    # Get the header row
    headers = [cell.value for cell in ws[1]]
    print(f"Existing columns: {headers}")

    # Find the column indices
    id_col = headers.index("id") + 1 if "id" in headers else 1
    name_col = headers.index("name") + 1 if "name" in headers else 2
    link_col = headers.index("link") + 1 if "link" in headers else 3
    ingredients_col = headers.index("ingredients") + 1 if "ingredients" in headers else 4
    allergensraw_col = headers.index("allergensraw") + 1 if "allergensraw" in headers else 5

    # Add new column header for allergensmapped
    new_col = len(headers) + 1

    # Style for header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add header for new column
    header_cell = ws.cell(row=1, column=new_col, value="allergensmapped")
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = header_alignment
    header_cell.border = thin_border

    # Process each row
    total_rows = ws.max_row
    mapped_count = 0

    print(f"\nProcessing {total_rows - 1} products...")

    for row in range(2, total_rows + 1):
        # Get the raw allergens
        allergensraw = ws.cell(row=row, column=allergensraw_col).value or ""

        # Also check ingredients for allergen mapping (more comprehensive)
        ingredients = ws.cell(row=row, column=ingredients_col).value or ""

        # Combine both for mapping
        combined_text = f"{allergensraw} {ingredients}"

        # Map to common allergens
        mapped = map_to_common_allergens(combined_text)

        # Write to new column
        cell = ws.cell(row=row, column=new_col, value=mapped)
        cell.border = thin_border

        if mapped:
            mapped_count += 1

        if row % 50 == 0:
            print(f"  Processed {row - 1}/{total_rows - 1} products...")

    # Set column width
    ws.column_dimensions[ws.cell(row=1, column=new_col).column_letter].width = 50

    # Save to new file
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")

    return total_rows - 1, mapped_count


def create_foodpreprocessed_from_scratch(input_file="foodraw.xlsx", output_file="foodpreprocessed.xlsx"):
    """
    Create a new preprocessed Excel file with all required columns
    """
    # Load source with data_only=True to get calculated values instead of formulas
    wb_src = load_workbook(input_file, data_only=True)
    ws_src = wb_src.active

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Food Data Preprocessed"

    # Define headers for the new file
    headers = ["id", "name", "link", "ingredients", "allergensraw", "allergensmapped"]

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Get source headers
    src_headers = [cell.value for cell in ws_src[1]]
    print(f"Source columns: {src_headers}")

    # Map source columns
    col_map = {}
    for idx, header in enumerate(src_headers):
        if header:
            col_map[header.lower()] = idx + 1

    # Prefer raw columns if available (ingredients_raw, allergens_raw)
    # These contain the original data without formulas
    ingredients_col = col_map.get('ingredients_raw', col_map.get('ingredients', 4))
    allergens_col = col_map.get('allergens_raw', col_map.get('allergensraw', 5))

    print(f"Reading ingredients from column {ingredients_col}")
    print(f"Reading allergens from column {allergens_col}")

    # Process each row
    total_rows = ws_src.max_row
    mapped_count = 0

    print(f"Processing {total_rows - 1} products...")

    for row in range(2, total_rows + 1):
        # Get source values
        id_val = ws_src.cell(row=row, column=col_map.get('id', 1)).value or ""
        name_val = ws_src.cell(row=row, column=col_map.get('name', 2)).value or ""
        link_val = ws_src.cell(row=row, column=col_map.get('link', 3)).value or ""

        # Read from raw columns to avoid formula issues
        ingredients_val = ws_src.cell(row=row, column=ingredients_col).value or ""
        allergensraw_val = ws_src.cell(row=row, column=allergens_col).value or ""

        # Skip if value looks like a formula (starts with =)
        if str(ingredients_val).startswith("="):
            ingredients_val = ""
        if str(allergensraw_val).startswith("="):
            allergensraw_val = ""

        # Map allergens
        combined_text = f"{allergensraw_val} {ingredients_val}"
        allergensmapped = map_to_common_allergens(combined_text)

        if allergensmapped:
            mapped_count += 1

        # Write to destination
        dest_row = row
        ws.cell(row=dest_row, column=1, value=id_val).border = thin_border
        ws.cell(row=dest_row, column=2, value=name_val).border = thin_border
        ws.cell(row=dest_row, column=3, value=link_val).border = thin_border
        ws.cell(row=dest_row, column=4, value=ingredients_val).border = thin_border
        ws.cell(row=dest_row, column=5, value=allergensraw_val).border = thin_border
        ws.cell(row=dest_row, column=6, value=allergensmapped).border = thin_border

        if row % 50 == 0:
            print(f"  Processed {row - 1}/{total_rows - 1} products...")

    # Set column widths
    ws.column_dimensions['A'].width = 8  # id
    ws.column_dimensions['B'].width = 40  # name
    ws.column_dimensions['C'].width = 55  # link
    ws.column_dimensions['D'].width = 80  # ingredients
    ws.column_dimensions['E'].width = 40  # allergensraw
    ws.column_dimensions['F'].width = 50  # allergensmapped

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(output_file)
    print(f"\nSaved to: {output_file}")

    return total_rows - 1, mapped_count


def main():
    print("=" * 60)
    print("Activity 03: Data Mapping")
    print("Mapping allergens to nine common allergens:")
    print("  1. Milk       4. Tree nut    7. Fish")
    print("  2. Egg        5. Wheat       8. Shellfish")
    print("  3. Peanut     6. Soy         9. Sesame")
    print("=" * 60)

    # Process the file
    total, mapped = create_foodpreprocessed_from_scratch(
        input_file="foodraw.xlsx",
        output_file="foodpreprocessed.xlsx"
    )

    # Summary statistics
    print("\n" + "=" * 60)
    print("Summary:")
    print(f"  Total products processed: {total}")
    print(f"  Products with mapped allergens: {mapped}")
    print(f"  Products without allergens: {total - mapped}")
    print("=" * 60)

    # Show allergen distribution
    print("\nAllergen Mapping Keywords:")
    for allergen, keywords in ALLERGEN_MAPPINGS.items():
        print(f"  {allergen.title()}: {', '.join(keywords[:5])}...")


if __name__ == "__main__":
    main()